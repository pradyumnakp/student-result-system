from openpyxl import Workbook

# ---------- INPUT ----------
def get_input(prompt, max_marks):
    while True:
        try:
            x = float(input(prompt))
            if 0 <= x <= max_marks:
                return x
            else:
                print("Enter between 0 and", max_marks)
        except:
            print("Invalid input")

# ---------- GRADE FUNCTION (AS PER IMAGE) ----------
def get_grade(score, max_marks):

    if max_marks == 10:
        if 9 <= score <= 10: return "A+"
        elif 7 <= score <= 8: return "A"
        elif 5 <= score <= 6: return "B+"
        elif 3 <= score <= 4: return "B"
        else: return "C"

    elif max_marks == 30:
        if 27 <= score <= 30: return "A+"
        elif 22 <= score <= 26: return "A"
        elif 16 <= score <= 21: return "B+"
        elif 10 <= score <= 15: return "B"
        else: return "C"

    elif max_marks == 50:
        if 45 <= score <= 50: return "A+"
        elif 35 <= score <= 44: return "A"
        elif 25 <= score <= 34: return "B+"
        elif 15 <= score <= 24: return "B"
        else: return "C"

    elif max_marks == 100:
        if 90 <= score <= 100: return "A+"
        elif 70 <= score <= 89: return "A"
        elif 50 <= score <= 69: return "B+"
        elif 30 <= score <= 49: return "B"
        else: return "C"

# ---------- FA ----------
def fa_input(name):
    t = get_input(f"{name} Test (25): ", 25)
    p1 = get_input("P1 (5): ", 5)
    p2 = get_input("P2 (5): ", 5)
    p3 = get_input("P3 (5): ", 5)
    total = t + p1 + p2 + p3
    red = (total / 40) * 10
    return total, red

# ---------- MAIN ----------
name = input("Student Name: ")
n = int(input("No. of Subjects: "))

subjects = []

for i in range(n):
    print(f"\nSubject {i+1}")
    sub_name = input("Subject name: ")

    # SEM 1
    fa1 = fa_input("FA1")
    fa2 = fa_input("FA2")

    sa1_w = get_input("SA1 Written (40): ", 40)
    sa1_o = get_input("SA1 Oral (10): ", 10)
    sa1_total = sa1_w + sa1_o
    sa1_red = (sa1_total / 50) * 30

    sem1 = fa1[1] + fa2[1] + sa1_red

    # SEM 2
    fa3 = fa_input("FA3")
    fa4 = fa_input("FA4")

    sa2_w = get_input("SA2 Written (40): ", 40)
    sa2_o = get_input("SA2 Oral (10): ", 10)
    sa2_total = sa2_w + sa2_o
    sa2_red = (sa2_total / 50) * 30

    sem2 = fa3[1] + fa4[1] + sa2_red

    final = sem1 + sem2

    subjects.append([
        sub_name,

        fa1[0], fa1[1], get_grade(fa1[1], 10),
        fa2[0], fa2[1], get_grade(fa2[1], 10),
        sa1_total, sa1_red, get_grade(sa1_red, 30),

        fa3[0], fa3[1], get_grade(fa3[1], 10),
        fa4[0], fa4[1], get_grade(fa4[1], 10),
        sa2_total, sa2_red, get_grade(sa2_red, 30),

        sem1, get_grade(sem1, 50),
        sem2, get_grade(sem2, 50),
        final, get_grade(final, 100)
    ])

# ---------- EXCEL ----------
wb = Workbook()
ws = wb.active

headers = [
    "Student","Subject",

    "FA1 Total","FA1 Red","Grade",
    "FA2 Total","FA2 Red","Grade",
    "SA1 Total","SA1 Red","Grade",

    "FA3 Total","FA3 Red","Grade",
    "FA4 Total","FA4 Red","Grade",
    "SA2 Total","SA2 Red","Grade",

    "Sem1","Grade",
    "Sem2","Grade",
    "Final","Grade"
]

ws.append(headers)

# ---------- WRITE ----------
first = True
for sub in subjects:
    row = []

    if first:
        row.append(name)
        first = False
    else:
        row.append("")

    row.extend(sub)
    ws.append(row)

# ---------- TOTAL MARKS ----------
total_marks = sum(sub[-2] for sub in subjects)   # sum of final

final_percentage = total_marks / len(subjects)

row_num = ws.max_row + 2

ws.cell(row=row_num, column=1, value="Total Marks:")
ws.cell(row=row_num, column=2, value=round(final_percentage, 2))
ws.cell(row=row_num, column=3, value=get_grade(final_percentage, 100))

# ---------- SAVE ----------
wb.save("report_format_like_image.xlsx")

print("\n✅ DONE: Excel with Grades + Total Marks created")
